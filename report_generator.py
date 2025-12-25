"""
Report Generator Module
Handles creating a comprehensive, multi-sheet Excel report.
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
import os

class ReportGenerator:
    def __init__(self):
        self.output_dir = "output"
        if not os.path.exists(self.output_dir):
            os.makedirs(self.output_dir)
        
        # Mapping of English column names to Arabic
        self.column_mapping = {
            # Employee Summary
            'Name': 'اسم الموظف',
            'Department': 'القسم',
            'Total Working Days': 'عدد أيام الدوام',
            'Complete Days': 'الأيام المستوفية',
            'Incomplete Days': 'أيام النقص',
            'Absent Days': 'أيام الغياب',
            'LateCount': 'عدد مرات التأخير',
            'LateMinutes': 'مجموع دقائق التأخير',
            'Actual Checks': 'البصمات الفعلية',
            'Required Checks': 'البصمات المطلوبة',
            'Missing Checks': 'البصمات المفقودة',
            'Compliance Rate': 'نسبة الالتزام %',
            # Daily Details
            'Date': 'التاريخ',
            'Day Status': 'الحالة اليومية',
            'Attendance Status': 'تفاصيل المطابقة',
            # Late Analysis
            'Required Time': 'الوقت المطلوب',
            'Actual Time': 'الوقت الفعلي',
            'Delay (minutes)': 'التأخير (بالدقائق)',
            # Matching Log
            'Tolerance Window': 'نافذة التسامح',
            'Match Result': 'نتيجة المطابقة',
            'Matched Punch': 'البصمة المطابقة',
            'Failure Reason': 'سبب الفشل'
        }

    def generate_full_report(self, employee_summary_data, daily_details_data, filename=None, tolerance_minutes=30):
        """
        Generate full attendance report with validation
        
        Args:
            employee_summary_data: DataFrame with employee summary
            daily_details_data: DataFrame with daily details
            filename: Output filename (optional)
            tolerance_minutes: Tolerance in minutes for late calculation (default: 30)
        """
        self.tolerance_minutes = tolerance_minutes
        # Validate employee summary data
        if employee_summary_data is None or employee_summary_data.empty:
            raise ValueError("بيانات ملخص الموظفين فارغة أو غير صحيحة")
        
        required_summary_cols = ['Name', 'Department', 'Total Working Days', 'Complete Days', 
                                'Incomplete Days', 'Absent Days', 'Actual Checks', 'Required Checks']
        missing_summary_cols = [col for col in required_summary_cols if col not in employee_summary_data.columns]
        if missing_summary_cols:
            raise ValueError(f"بيانات ملخص الموظفين غير مكتملة. أعمدة مفقودة: {', '.join(missing_summary_cols)}")
        
        # Validate daily details data
        if daily_details_data is None or daily_details_data.empty:
            raise ValueError("بيانات التفاصيل اليومية فارغة أو غير صحيحة")
        
        required_daily_cols = ['Name', 'Department', 'Date', 'Day Status', 'Actual Checks', 'Required Checks']
        missing_daily_cols = [col for col in required_daily_cols if col not in daily_details_data.columns]
        if missing_daily_cols:
            raise ValueError(f"بيانات التفاصيل اليومية غير مكتملة. أعمدة مفقودة: {', '.join(missing_daily_cols)}")
        
        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"full_attendance_report_{timestamp}.xlsx"
        
        filepath = os.path.join(self.output_dir, filename)
        
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            # Sheet 1: Employee Summary
            self._write_employee_summary_sheet(writer, employee_summary_data)
            
            # Sheet 2: Daily Attendance Details
            self._write_daily_details_sheet(writer, daily_details_data)
            
            # Sheet 3: Daily Attendance Sheet (تفاصيل الحضور اليومي)
            self._write_daily_attendance_sheet(writer, daily_details_data)
            
            # Sheet 4: Absences (الغيابات)
            self._write_absences_sheet(writer, employee_summary_data)
            
            # Sheet 5: Missing Punches (البصمات المتروكة)
            self._write_missing_punches_sheet(writer, daily_details_data)
            
            # Sheet 6: Raw Fingerprint Matching Log
            self._write_matching_log_sheet(writer, daily_details_data)
        
        return filepath

    def _write_employee_summary_sheet(self, writer, data):
        # Create a copy and rename columns to Arabic
        data_arabic = data.copy()
        data_arabic.columns = [self.column_mapping.get(col, col) for col in data_arabic.columns]
        
        data_arabic.to_excel(writer, sheet_name='ملخص الموظفين', index=False)
        ws = writer.sheets['ملخص الموظفين']
        self._format_sheet(ws, {
            'أيام الغياب': 'FFC7CE',      # Light Red
            'أيام النقص': 'FFEB9C', # Yellow
            'الأيام المستوفية': 'C6EFCE',    # Light Green
        })

    def _write_daily_details_sheet(self, writer, data):
        """
        Create daily details sheet - one row per day
        Missing punches details are available in the separate "البصمات المتروكة" sheet
        """
        # Select columns that exist in data
        available_cols = ['Name', 'Department', 'Date', 'Day Status', 'LateCount', 'LateMinutes', 
                         'Actual Checks', 'Required Checks', 'Missing Checks', 'Compliance Rate']
        cols_to_use = [col for col in available_cols if col in data.columns]
        daily_data = data[cols_to_use].copy()
        
        # Rename columns to Arabic
        daily_data.columns = [self.column_mapping.get(col, col) for col in daily_data.columns]
        
        # Sort by employee name, then by date
        if 'اسم الموظف' in daily_data.columns and 'التاريخ' in daily_data.columns:
            daily_data = daily_data.sort_values(['اسم الموظف', 'التاريخ'])
        
        daily_data.to_excel(writer, sheet_name='تفاصيل الحضور اليومي', index=False)
        ws = writer.sheets['تفاصيل الحضور اليومي']
        
        # Format sheet with dynamic status values
        self._format_sheet(ws, {
            'الحالة اليومية': {
                'غائب': 'FFC7CE',
                'مستوفي': 'C6EFCE',
            }
        })
        
        # Apply yellow background for rows with "نقص بصمة" in status
        from openpyxl.styles import PatternFill
        yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        
        # Find the status column index
        status_col_idx = None
        for idx, header in enumerate(ws[1], 1):
            if header.value == 'الحالة اليومية':
                status_col_idx = idx
                break
        
        # Apply yellow to rows with "نقص بصمة" in status
        if status_col_idx:
            for row_idx in range(2, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=status_col_idx)
                if cell.value and 'نقص بصمة' in str(cell.value):
                    cell.fill = yellow_fill

    def _write_late_analysis_sheet(self, writer, data):
        """
        Create late analysis sheet showing all late arrivals
        Only includes delays that exceed tolerance_minutes
        """
        late_data = []
        
        # Check if data has required columns
        if data is None or data.empty:
            # Create empty sheet with headers
            empty_df = pd.DataFrame(columns=['اسم الموظف', 'التاريخ', 'الوقت المطلوب', 'الوقت الفعلي', 'التأخير (بالدقائق)'])
            empty_df.to_excel(writer, sheet_name='تحليل التأخير', index=False)
            ws = writer.sheets['تحليل التأخير']
            self._format_sheet(ws, {})
            return
        
        # Check if Attendance Status column exists
        if 'Attendance Status' not in data.columns:
            # Try to get from original data if it was filtered
            # For now, create empty sheet
            empty_df = pd.DataFrame(columns=['اسم الموظف', 'التاريخ', 'الوقت المطلوب', 'الوقت الفعلي', 'التأخير (بالدقائق)'])
            empty_df.to_excel(writer, sheet_name='تحليل التأخير', index=False)
            ws = writer.sheets['تحليل التأخير']
            self._format_sheet(ws, {})
            return
        
        for _, row in data.iterrows():
            # Check if LateCount exists and is greater than 0
            late_count = row.get('LateCount', 0)
            try:
                late_count = float(late_count) if not pd.isna(late_count) else 0
            except (ValueError, TypeError):
                late_count = 0
                
            if late_count <= 0:
                continue
                
            # Check if Attendance Status exists and is a list
            attendance_status = row.get('Attendance Status', [])
            
            # Handle case where Attendance Status might be stored as string
            if isinstance(attendance_status, str):
                try:
                    import ast
                    attendance_status = ast.literal_eval(attendance_status)
                except:
                    continue
                    
            if not isinstance(attendance_status, list) or len(attendance_status) == 0:
                continue
                
            for status in attendance_status:
                if not isinstance(status, dict):
                    continue
                
                actual_time = status.get('actual_time')
                required_time = status.get('required_time')
                if actual_time is None or required_time is None:
                    continue
                
                # Include if:
                # 1. Matched within tolerance and delay > tolerance_minutes, OR
                # 2. Exceeded tolerance (outside tolerance window)
                matched = status.get('matched', False)
                exceeded_tolerance = status.get('exceeded_tolerance', False)
                
                if not matched and not exceeded_tolerance:
                    continue  # Skip if no fingerprint found
                    
                try:
                    # Calculate delay in minutes
                    if hasattr(actual_time, 'strftime') and hasattr(required_time, 'strftime'):
                        delay = (actual_time - required_time).total_seconds() / 60
                    else:
                        # Try to convert to datetime if they're strings
                        try:
                            from datetime import datetime
                            if isinstance(actual_time, str):
                                # Try different formats
                                for fmt in ['%H:%M:%S', '%H:%M', '%Y-%m-%d %H:%M:%S']:
                                    try:
                                        actual_time = datetime.strptime(actual_time, fmt)
                                        break
                                    except:
                                        continue
                            if isinstance(required_time, str):
                                for fmt in ['%H:%M:%S', '%H:%M', '%Y-%m-%d %H:%M:%S']:
                                    try:
                                        required_time = datetime.strptime(required_time, fmt)
                                        break
                                    except:
                                        continue
                            if hasattr(actual_time, 'strftime') and hasattr(required_time, 'strftime'):
                                delay = (actual_time - required_time).total_seconds() / 60
                            else:
                                continue
                        except:
                            continue
                    
                    # Count as late if:
                    # 1. Matched within tolerance but delay > tolerance_minutes, OR
                    # 2. Exceeded tolerance (outside tolerance window)
                    if exceeded_tolerance or (matched and delay > self.tolerance_minutes):
                        late_minutes = int(delay - self.tolerance_minutes)  # Only count minutes beyond tolerance
                        late_data.append({
                            'اسم الموظف': row.get('Name', 'N/A'),
                            'التاريخ': row.get('Date', 'N/A'),
                            'الوقت المطلوب': status.get('required_time_str', 'N/A'),
                            'الوقت الفعلي': actual_time.strftime('%H:%M:%S') if hasattr(actual_time, 'strftime') else str(actual_time),
                            'التأخير (بالدقائق)': late_minutes
                        })
                except (AttributeError, TypeError, ValueError) as e:
                    # Skip this entry if there's an error calculating delay
                    continue
        
        late_df = pd.DataFrame(late_data)
        if not late_df.empty:
            late_df.to_excel(writer, sheet_name='تحليل التأخير', index=False)
            ws = writer.sheets['تحليل التأخير']
            self._format_sheet(ws, {})
        else:
            # Create empty sheet with headers
            empty_df = pd.DataFrame(columns=['اسم الموظف', 'التاريخ', 'الوقت المطلوب', 'الوقت الفعلي', 'التأخير (بالدقائق)'])
            empty_df.to_excel(writer, sheet_name='تحليل التأخير', index=False)
            ws = writer.sheets['تحليل التأخير']
            self._format_sheet(ws, {})

    def _write_matching_log_sheet(self, writer, data):
        log_data = []
        for _, row in data.iterrows():
            # Check if Attendance Status exists and is a list
            attendance_status = row.get('Attendance Status', [])
            if not isinstance(attendance_status, list):
                continue
                
            for status in attendance_status:
                if not isinstance(status, dict):
                    continue
                    
                # Safely get values with defaults
                tolerance_start = status.get('tolerance_start')
                tolerance_end = status.get('tolerance_end')
                actual_time = status.get('actual_time')
                matched = status.get('matched', False)
                exceeded_tolerance = status.get('exceeded_tolerance', False)
                required_time = status.get('required_time')
                
                # Format tolerance window safely
                if tolerance_start is not None and hasattr(tolerance_start, 'strftime'):
                    tolerance_start_str = tolerance_start.strftime('%H:%M')
                else:
                    tolerance_start_str = 'N/A'
                    
                if tolerance_end is not None and hasattr(tolerance_end, 'strftime'):
                    tolerance_end_str = tolerance_end.strftime('%H:%M')
                else:
                    tolerance_end_str = 'N/A'
                
                # Format actual time safely
                actual_time_str = '—'
                delay_minutes = None
                if actual_time is not None and hasattr(actual_time, 'strftime'):
                    actual_time_str = actual_time.strftime('%H:%M')
                    if required_time and hasattr(required_time, 'strftime'):
                        delay_seconds = (actual_time - required_time).total_seconds()
                        delay_minutes = int(delay_seconds / 60)
                
                # Determine match status
                if exceeded_tolerance:
                    match_status = 'تجاوز نافذة السماح'
                elif matched:
                    match_status = 'مطابق'
                else:
                    match_status = 'غير مطابق'
                
                log_data.append({
                    'اسم الموظف': row['Name'],
                    'التاريخ': row['Date'],
                    'الوقت المطلوب': status.get('required_time_str', 'N/A'),
                    'الوقت الفعلي': actual_time_str,
                    'التأخير (دقيقة)': delay_minutes if delay_minutes is not None else '—',
                    'نافذة التسامح': f"{tolerance_start_str} - {tolerance_end_str}",
                    'الحالة': match_status
                })
        
        log_df = pd.DataFrame(log_data)
        if not log_df.empty:
            # Reorder columns for better readability
            column_order = ['اسم الموظف', 'التاريخ', 'الوقت المطلوب', 'الوقت الفعلي', 
                          'التأخير (دقيقة)', 'نافذة التسامح', 'الحالة']
            log_df = log_df[column_order]
            
            log_df.to_excel(writer, sheet_name='سجل مطابقة البصمات', index=False)
            ws = writer.sheets['سجل مطابقة البصمات']
            self._format_matching_log_sheet(ws)
        else:
            # Create empty sheet with headers
            empty_df = pd.DataFrame(columns=['اسم الموظف', 'التاريخ', 'الوقت المطلوب', 'الوقت الفعلي', 
                                            'التأخير (دقيقة)', 'نافذة التسامح', 'الحالة'])
            empty_df.to_excel(writer, sheet_name='سجل مطابقة البصمات', index=False)
            ws = writer.sheets['سجل مطابقة البصمات']
            self._format_matching_log_sheet(ws)
    
    def _write_missing_punches_sheet(self, writer, data):
        """
        Create a sheet showing missing punches sorted by employee name
        """
        missing_punches_data = []
        
        for _, row in data.iterrows():
            # Check if Attendance Status exists and is a list
            attendance_status = row.get('Attendance Status', [])
            if not isinstance(attendance_status, list):
                continue
            
            # Get employee info
            emp_name = row.get('Name', 'N/A')
            emp_dept = row.get('Department', 'N/A')
            date = row.get('Date', 'N/A')
            
            # Find all unmatched required times (missing punches)
            for status in attendance_status:
                if not isinstance(status, dict):
                    continue
                
                matched = status.get('matched', False)
                if not matched:  # This is a missing punch
                    required_time_str = status.get('required_time_str', 'N/A')
                    tolerance_start = status.get('tolerance_start')
                    tolerance_end = status.get('tolerance_end')
                    
                    # Format tolerance window
                    if tolerance_start is not None and hasattr(tolerance_start, 'strftime'):
                        tolerance_start_str = tolerance_start.strftime('%H:%M')
                    else:
                        tolerance_start_str = 'N/A'
                        
                    if tolerance_end is not None and hasattr(tolerance_end, 'strftime'):
                        tolerance_end_str = tolerance_end.strftime('%H:%M')
                    else:
                        tolerance_end_str = 'N/A'
                    
                    tolerance_window = f"{tolerance_start_str} - {tolerance_end_str}"
                    
                    missing_punches_data.append({
                        'اسم الموظف': emp_name,
                        'القسم': emp_dept,
                        'التاريخ': date,
                        'الوقت المطلوب': required_time_str,
                        'نافذة التسامح': tolerance_window,
                        'الحالة': 'لم يبصم'
                    })
        
        # Create DataFrame and sort by employee name
        missing_df = pd.DataFrame(missing_punches_data)
        
        if not missing_df.empty:
            # Sort by employee name, then by date, then by required time
            missing_df = missing_df.sort_values(['اسم الموظف', 'التاريخ', 'الوقت المطلوب'])
            missing_df.to_excel(writer, sheet_name='البصمات المتروكة', index=False)
            ws = writer.sheets['البصمات المتروكة']
            self._format_sheet(ws, {})
        else:
            # Create empty sheet with headers
            empty_df = pd.DataFrame(columns=['اسم الموظف', 'القسم', 'التاريخ', 'الوقت المطلوب', 
                                            'نافذة التسامح', 'الحالة'])
            empty_df.to_excel(writer, sheet_name='البصمات المتروكة', index=False)
            ws = writer.sheets['البصمات المتروكة']
            self._format_sheet(ws, {})
    
    def _write_daily_attendance_sheet(self, writer, data):
        """
        Create detailed daily attendance sheet with time columns
        Each row represents one employee + one day
        Columns: اسم الموظف، التاريخ، 08:00، 12:00، 15:00، 20:00، 23:00، 08:00 (اليوم التالي)،
                 تأخير، تارك بصمة، قبل، ملاحظة
        """
        import ast
        
        sheet_data = []
        
        # Process each row in daily_details_data
        for _, row in data.iterrows():
            emp_name = row.get('Name', 'N/A')
            date = row.get('Date', 'N/A')
            day_status = row.get('Day Status', 'N/A')
            late_count = row.get('LateCount', 0)
            missing_checks = row.get('Missing Checks', 0)
            
            # Get attendance status
            attendance_status = row.get('Attendance Status', [])
            
            # Handle case where Attendance Status might be stored as string
            if isinstance(attendance_status, str):
                try:
                    attendance_status = ast.literal_eval(attendance_status)
                except:
                    attendance_status = []
            
            if not isinstance(attendance_status, list):
                attendance_status = []
            
            # Convert date to datetime for comparison
            if hasattr(date, 'date'):
                shift_date = date.date() if hasattr(date, 'date') else date
            elif hasattr(date, 'strftime'):
                try:
                    shift_date = datetime.strptime(str(date), '%Y-%m-%d').date()
                except:
                    shift_date = date
            else:
                try:
                    shift_date = datetime.strptime(str(date), '%Y-%m-%d').date()
                except:
                    shift_date = date
            
            # Create a dictionary to store time punches
            # Key: (req_time_str, is_next_day) to handle duplicate "08:00"
            time_punches = {}
            early_count = 0  # Count early arrivals (before required time)
            
            # Process attendance status and determine order
            status_list_with_index = []
            for idx, status in enumerate(attendance_status):
                if not isinstance(status, dict):
                    continue
                
                req_time_str = status.get('required_time_str', '')
                required_time = status.get('required_time')
                matched = status.get('matched', False)
                actual_time = status.get('actual_time')
                
                # Check if this is next day (by comparing date of required_time with shift_date)
                is_next_day = False
                if required_time and hasattr(required_time, 'date'):
                    req_date = required_time.date()
                    if isinstance(shift_date, type(req_date)):
                        is_next_day = req_date > shift_date
                
                # Create key: use index to distinguish duplicate time strings
                # The last one in the list is typically the next day one
                key = (req_time_str, idx, is_next_day)
                status_list_with_index.append((key, status, req_time_str, is_next_day))
                
                # Check if exceeded tolerance (outside tolerance window)
                exceeded_tolerance = status.get('exceeded_tolerance', False)
                
                if actual_time and required_time:
                    # Format actual time as HH:MM
                    if hasattr(actual_time, 'strftime'):
                        time_str = actual_time.strftime('%H:%M')
                    else:
                        time_str = str(actual_time)
                    
                    # Calculate delay
                    delay = (actual_time - required_time).total_seconds() / 60
                    is_late = delay > self.tolerance_minutes
                    is_early = delay < -self.tolerance_minutes
                    
                    if is_early:
                        early_count += 1
                    
                    time_punches[key] = {
                        'time': time_str,
                        'matched': matched,  # True only if within tolerance
                        'is_late': is_late,
                        'is_early': is_early,
                        'exceeded_tolerance': exceeded_tolerance,  # True if outside tolerance window
                        'delay': delay,
                        'req_time_str': req_time_str,
                        'is_next_day': is_next_day,
                        'index': idx
                    }
                else:
                    # Missing punch
                    time_punches[key] = {
                        'time': '—',
                        'matched': False,
                        'is_late': False,
                        'is_early': False,
                        'exceeded_tolerance': False,
                        'delay': None,
                        'req_time_str': req_time_str,
                        'is_next_day': is_next_day,
                        'index': idx
                    }
            
            # Build row data
            row_data = {
                'اسم الموظف': emp_name,
                'التاريخ': date.strftime('%Y-%m-%d') if hasattr(date, 'strftime') else str(date),
            }
            
            # Map time columns - use the order from attendance_status
            # Expected columns: 08:00, 12:00, 15:00, 20:00, 23:00, 08:00 (اليوم التالي)
            column_mapping = []
            for key, status, req_time_str, is_next_day in status_list_with_index:
                if req_time_str == '08:00' and is_next_day:
                    col_name = '08:00 (اليوم التالي)'
                else:
                    col_name = req_time_str
                column_mapping.append((col_name, key))
            
            # Fill time columns based on attendance_status order
            # Map each status to its column name, avoiding duplicates
            filled_columns = set()
            
            # First, handle next day 08:00 (should be last)
            next_day_08_found = False
            for key, status, req_time_str, is_next_day in status_list_with_index:
                if req_time_str == '08:00' and is_next_day:
                    col_name = '08:00 (اليوم التالي)'
                    if col_name not in filled_columns:
                        if key in time_punches:
                            row_data[col_name] = time_punches[key]['time']
                        else:
                            row_data[col_name] = '—'
                        filled_columns.add(col_name)
                        next_day_08_found = True
                    break
            
            # Then, handle regular time columns (in order, but skip next day 08:00)
            for key, status, req_time_str, is_next_day in status_list_with_index:
                if req_time_str == '08:00' and is_next_day:
                    continue  # Skip, already handled
                
                col_name = req_time_str
                if col_name not in filled_columns:
                    if key in time_punches:
                        row_data[col_name] = time_punches[key]['time']
                    else:
                        row_data[col_name] = '—'
                    filled_columns.add(col_name)
            
            # Ensure all expected columns exist (fill missing ones with —)
            expected_time_columns = ['08:00', '12:00', '15:00', '20:00', '23:00', '08:00 (اليوم التالي)']
            for col_name in expected_time_columns:
                if col_name not in row_data:
                    row_data[col_name] = '—'
            
            # Add status columns
            row_data['تأخير'] = late_count
            row_data['تارك بصمة'] = missing_checks
            row_data['قبل'] = 1 if early_count > 0 else 0
            row_data['ملاحظة'] = day_status  # Use day status as note
            
            sheet_data.append(row_data)
        
        # Create DataFrame
        df = pd.DataFrame(sheet_data)
        
        # Ensure all columns are present (in case some are missing)
        expected_columns = ['اسم الموظف', 'التاريخ', '08:00', '12:00', '15:00', '20:00', 
                          '23:00', '08:00 (اليوم التالي)', 'تأخير', 'تارك بصمة', 'قبل', 'ملاحظة']
        for col in expected_columns:
            if col not in df.columns:
                df[col] = ''
        
        # Reorder columns to match expected order
        df = df[expected_columns]
        
        # Sort: التاريخ (تصاعدي), اسم الموظف (أبجدي), الحالة (مستوفي → نقص بصمة → غياب)
        # Create sort key for status
        def get_status_sort_key(status_str):
            if not isinstance(status_str, str):
                return 3
            if status_str == 'مستوفي':
                return 0
            elif status_str.startswith('نقص بصمة'):
                return 1
            elif status_str == 'غائب':
                return 2
            else:
                return 3
        
        df['_status_sort'] = df['ملاحظة'].apply(get_status_sort_key)
        df = df.sort_values(['التاريخ', 'اسم الموظف', '_status_sort'])
        df = df.drop('_status_sort', axis=1)
        
        # Write to Excel
        if not df.empty:
            df.to_excel(writer, sheet_name='تفاصيل الحضور اليومي', index=False)
            ws = writer.sheets['تفاصيل الحضور اليومي']
            
            # Format sheet with colors
            self._format_daily_attendance_sheet(ws, data)
        else:
            # Create empty sheet with headers
            empty_df = pd.DataFrame(columns=expected_columns)
            empty_df.to_excel(writer, sheet_name='تفاصيل الحضور اليومي', index=False)
            ws = writer.sheets['تفاصيل الحضور اليومي']
            self._format_daily_attendance_sheet(ws, pd.DataFrame())
    
    def _format_daily_attendance_sheet(self, ws, original_data):
        """
        Format the daily attendance sheet with colors and styling
        """
        import ast
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        # Header formatting
        header_font = Font(bold=True, name='Arial', size=11, color="FFFFFF")
        header_fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")  # Gray background
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Format header row
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Get column indices
        col_indices = {}
        for idx, cell in enumerate(ws[1], 1):
            col_indices[cell.value] = idx
        
        # Define time column names (in order)
        time_columns = ['08:00', '12:00', '15:00', '20:00', '23:00', '08:00 (اليوم التالي)']
        
        # Color definitions
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Light green
        late_red_fill = PatternFill(start_color="DC143C", end_color="DC143C", fill_type="solid")  # Crimson red for late punches (within tolerance but late)
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Light red for missing punches
        bordo_fill = PatternFill(start_color="800000", end_color="800000", fill_type="solid")  # Bordo/Dark red for exceeded tolerance (outside tolerance window)
        gray_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")  # Gray
        
        # Process each data row (starting from row 2)
        for row_idx in range(2, ws.max_row + 1):
            # Get employee name and date to find original data
            emp_name = ws.cell(row=row_idx, column=col_indices.get('اسم الموظف', 1)).value
            date_str = ws.cell(row=row_idx, column=col_indices.get('التاريخ', 2)).value
            
            # Find corresponding row in original data
            attendance_status_list = []
            if not original_data.empty:
                matching_rows = original_data[
                    (original_data['Name'] == emp_name) & 
                    (original_data['Date'].astype(str) == str(date_str))
                ]
                if not matching_rows.empty:
                    att_status = matching_rows.iloc[0].get('Attendance Status', [])
                    if isinstance(att_status, str):
                        try:
                            attendance_status_list = ast.literal_eval(att_status)
                        except:
                            attendance_status_list = []
                    elif isinstance(att_status, list):
                        attendance_status_list = att_status
            
            # Get shift date for comparison
            try:
                shift_date = datetime.strptime(str(date_str), '%Y-%m-%d').date()
            except:
                shift_date = None
            
            # Create map of (req_time_str, is_next_day, index) to status info
            time_status_map = {}
            for idx, status in enumerate(attendance_status_list):
                if isinstance(status, dict):
                    req_time_str = status.get('required_time_str', '')
                    matched = status.get('matched', False)
                    actual_time = status.get('actual_time')
                    required_time = status.get('required_time')
                    
                    # Check if this is next day
                    is_next_day = False
                    if required_time and shift_date and hasattr(required_time, 'date'):
                        req_date = required_time.date()
                        is_next_day = req_date > shift_date
                    
                    key = (req_time_str, idx, is_next_day)
                    
                    if matched and actual_time and required_time:
                        delay = (actual_time - required_time).total_seconds() / 60
                        is_late = delay > self.tolerance_minutes
                        exceeded_tolerance = status.get('exceeded_tolerance', False)
                        time_status_map[key] = {
                            'matched': True, 
                            'is_late': is_late, 
                            'exceeded_tolerance': exceeded_tolerance,
                            'req_time_str': req_time_str, 
                            'is_next_day': is_next_day
                        }
                    elif actual_time and required_time:
                        # Has actual_time but not matched (exceeded tolerance)
                        delay = (actual_time - required_time).total_seconds() / 60
                        exceeded_tolerance = status.get('exceeded_tolerance', False)
                        time_status_map[key] = {
                            'matched': False, 
                            'is_late': True,  # Always late if exceeded tolerance
                            'exceeded_tolerance': exceeded_tolerance,
                            'req_time_str': req_time_str, 
                            'is_next_day': is_next_day
                        }
                    else:
                        time_status_map[key] = {
                            'matched': False, 
                            'is_late': False, 
                            'exceeded_tolerance': False,
                            'req_time_str': req_time_str, 
                            'is_next_day': is_next_day
                        }
            
            # Color time columns
            for time_col in time_columns:
                if time_col in col_indices:
                    col_idx = col_indices[time_col]
                    cell = ws.cell(row=row_idx, column=col_idx)
                    
                    cell_value = cell.value
                    is_next_day_col = time_col == '08:00 (اليوم التالي)'
                    
                    if cell_value and cell_value != '—':
                        # Has fingerprint - check if late or on time
                        # Find matching status (handle next day case)
                        found_status = None
                        for key, status_info in time_status_map.items():
                            req_time_str, idx, is_next_day = key
                            if is_next_day_col:
                                if status_info['req_time_str'] == '08:00' and status_info['is_next_day']:
                                    found_status = status_info
                                    break
                            else:
                                if status_info['req_time_str'] == time_col and not status_info['is_next_day']:
                                    found_status = status_info
                                    break
                        
                        if found_status:
                            # Check if exceeded tolerance first (highest priority)
                            if found_status.get('exceeded_tolerance', False):
                                cell.fill = bordo_fill  # Bordo: exceeded tolerance (outside tolerance window)
                            elif found_status['matched']:
                                if found_status['is_late']:
                                    cell.fill = late_red_fill  # Crimson red: late but within tolerance window
                                else:
                                    cell.fill = green_fill  # Green: on time
                            else:
                                # Has actual_time but not matched (should not happen normally)
                                cell.fill = bordo_fill  # Bordo: exceeded tolerance
                        else:
                            cell.fill = green_fill  # Default to green if status not found
                    else:
                        # Missing fingerprint - red
                        cell.fill = red_fill
            
            # Format status columns
            # تأخير column
            if 'تأخير' in col_indices:
                delay_col = col_indices['تأخير']
                delay_cell = ws.cell(row=row_idx, column=delay_col)
                delay_value = delay_cell.value
                try:
                    delay_int = int(delay_value) if delay_value is not None else 0
                    if delay_int == 0:
                        delay_cell.fill = green_fill
                    else:
                        delay_cell.fill = late_red_fill  # Crimson red for late punches
                except:
                    pass
            
            # تارك بصمة column
            if 'تارك بصمة' in col_indices:
                missing_col = col_indices['تارك بصمة']
                missing_cell = ws.cell(row=row_idx, column=missing_col)
                missing_value = missing_cell.value
                try:
                    missing_int = int(missing_value) if missing_value is not None else 0
                    if missing_int == 0:
                        missing_cell.fill = green_fill
                    else:
                        missing_cell.fill = red_fill
                except:
                    pass
            
            # قبل column
            if 'قبل' in col_indices:
                early_col = col_indices['قبل']
                early_cell = ws.cell(row=row_idx, column=early_col)
                early_value = early_cell.value
                try:
                    early_int = int(early_value) if early_value is not None else 0
                    if early_int == 1:
                        early_cell.fill = green_fill
                    else:
                        early_cell.fill = gray_fill
                except:
                    pass
        
        # Set column widths
        column_widths = {
            'اسم الموظف': 20,
            'التاريخ': 12,
            '08:00': 10,
            '12:00': 10,
            '15:00': 10,
            '20:00': 10,
            '23:00': 10,
            '08:00 (اليوم التالي)': 18,
            'تأخير': 10,
            'تارك بصمة': 12,
            'قبل': 8,
            'ملاحظة': 20
        }
        
        for col_name, width in column_widths.items():
            if col_name in col_indices:
                col_letter = ws.cell(row=1, column=col_indices[col_name]).column_letter
                ws.column_dimensions[col_letter].width = width
        
        # Set RTL alignment for all cells
        cell_alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                if cell.alignment.horizontal is None or cell.alignment.horizontal == 'general':
                    cell.alignment = cell_alignment
        
        # Enable filters
        ws.auto_filter.ref = ws.dimensions

    def _write_absences_sheet(self, writer, data):
        """
        Create a sheet showing absences with reasons, sorted by employee name
        """
        absences_data = []
        
        for _, row in data.iterrows():
            absent_days = row.get('Absent Days', 0)
            if absent_days > 0:
                absences_data.append({
                    'اسم الموظف': row.get('Name', 'N/A'),
                    'القسم': row.get('Department', 'N/A'),
                    'عدد أيام الغياب': absent_days,
                    'عدد البصمات المتروكة': row.get('Missing Checks', 0),
                    'عدد أيام الدوام': row.get('Total Working Days', 0),
                    'نسبة الالتزام %': row.get('Compliance Rate', 0)
                })
        
        # Create DataFrame and sort by employee name
        absences_df = pd.DataFrame(absences_data)
        
        if not absences_df.empty:
            # Sort by employee name
            absences_df = absences_df.sort_values('اسم الموظف')
            absences_df.to_excel(writer, sheet_name='الغيابات', index=False)
            ws = writer.sheets['الغيابات']
            self._format_sheet(ws, {
                'عدد أيام الغياب': 'FFC7CE',  # Light Red for absence days
            })
        else:
            # Create empty sheet with headers
            empty_df = pd.DataFrame(columns=['اسم الموظف', 'القسم', 'عدد أيام الغياب', 
                                            'عدد البصمات المتروكة', 
                                            'عدد أيام الدوام', 'نسبة الالتزام %'])
            empty_df.to_excel(writer, sheet_name='الغيابات', index=False)
            ws = writer.sheets['الغيابات']
            self._format_sheet(ws, {})

    def _format_matching_log_sheet(self, ws):
        """
        Format the matching log sheet with professional styling
        """
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        # Header formatting - professional blue gradient
        header_font = Font(bold=True, name='Arial', size=11, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")  # Professional blue
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Format header row
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Get column indices
        col_indices = {}
        for idx, cell in enumerate(ws[1], 1):
            col_indices[cell.value] = idx
        
        # Color definitions
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Light green for matched
        late_red_fill = PatternFill(start_color="DC143C", end_color="DC143C", fill_type="solid")  # Crimson red for late
        bordo_fill = PatternFill(start_color="800000", end_color="800000", fill_type="solid")  # Bordo for exceeded tolerance
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Light red for not matched
        
        # Format data rows
        for row_idx in range(2, ws.max_row + 1):
            # Get status value
            status_col_idx = col_indices.get('الحالة')
            if status_col_idx:
                status_cell = ws.cell(row=row_idx, column=status_col_idx)
                status_value = status_cell.value
                
                # Color code based on status
                if status_value == 'مطابق':
                    status_cell.fill = green_fill
                elif status_value == 'تجاوز نافذة السماح':
                    status_cell.fill = bordo_fill
                    # Also color the actual time cell
                    if 'الوقت الفعلي' in col_indices:
                        actual_time_cell = ws.cell(row=row_idx, column=col_indices['الوقت الفعلي'])
                        actual_time_cell.fill = bordo_fill
                elif status_value == 'غير مطابق':
                    status_cell.fill = red_fill
            
            # Color delay column if late
            if 'التأخير (دقيقة)' in col_indices:
                delay_col = col_indices['التأخير (دقيقة)']
                delay_cell = ws.cell(row=row_idx, column=delay_col)
                delay_value = delay_cell.value
                if delay_value and delay_value != '—':
                    try:
                        delay_int = int(delay_value) if delay_value is not None else 0
                        if delay_int > 0:
                            delay_cell.fill = late_red_fill
                    except:
                        pass
        
        # Set column widths for optimal readability
        column_widths = {
            'اسم الموظف': 20,
            'التاريخ': 12,
            'الوقت المطلوب': 15,
            'الوقت الفعلي': 15,
            'التأخير (دقيقة)': 18,
            'نافذة التسامح': 20,
            'الحالة': 20
        }
        
        for col_name, width in column_widths.items():
            if col_name in col_indices:
                col_letter = ws.cell(row=1, column=col_indices[col_name]).column_letter
                ws.column_dimensions[col_letter].width = width
        
        # Set RTL alignment for all cells
        cell_alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                if cell.alignment.horizontal is None or cell.alignment.horizontal == 'general':
                    cell.alignment = cell_alignment
        
        # Add borders for better readability
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = thin_border
        
        # Enable filters
        ws.auto_filter.ref = ws.dimensions
        
        # Freeze first row
        ws.freeze_panes = 'A2'

    def _format_sheet(self, ws, color_map):
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for col_idx, column in enumerate(ws.columns):
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 40)
            ws.column_dimensions[column_letter].width = adjusted_width

            # Apply conditional formatting
            if column[0].value in color_map:
                rules = color_map[column[0].value]
                if isinstance(rules, dict): # Rule per value
                    for cell in column[1:]:
                        if cell.value in rules:
                            cell.fill = PatternFill(start_color=rules[cell.value], end_color=rules[cell.value], fill_type="solid")
                else: # Rule for any non-zero value
                    color = rules
                    for cell in column[1:]:
                        if cell.value and int(cell.value) > 0:
                             cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")


def main():
    """
    Example usage of the ReportGenerator
    """
    # Create dummy data for testing
    employee_summary = pd.DataFrame({
        'Name': ['Ahmad', 'Fatima'],
        'Department': ['Admin', 'Accounting'],
        'Complete Days': [20, 18],
        'Incomplete Days': [2, 3],
        'Absent Days': [0, 1],
    })

    daily_details = pd.DataFrame({
        'Name': ['Ahmad', 'Ahmad', 'Fatima'],
        'Department': ['Admin', 'Admin', 'Accounting'],
        'Date': [datetime(2023, 1, 1), datetime(2023, 1, 2), datetime(2023, 1, 1)],
        'Day Status': ['مستوفي', 'نقص بصمة', 'غائب'],
        'LateCount': [0, 1, 0],
        'LateMinutes': [0, 15, 0],
        'Attendance Status': [[{'matched': True, 'required_time': datetime(2023,1,1,8,0), 'required_time_str': '08:00', 'actual_time': datetime(2023,1,1,8,5), 'tolerance_start': datetime(2023,1,1,7,30), 'tolerance_end': datetime(2023,1,1,8,30)}]] * 3
    })
    
    generator = ReportGenerator()
    filepath = generator.generate_full_report(employee_summary, daily_details)
    print(f"Full report generated at: {filepath}")

if __name__ == '__main__':
    main()